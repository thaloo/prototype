import csv
import os.path
import codecs
import json
from django.shortcuts import render
from django.shortcuts import render_to_response
from django.shortcuts import HttpResponse
from urllib.request import urlopen
from urllib.request import Request
from urllib.request import urlretrieve
from urllib.parse import quote
from urllib.parse import unquote
from bs4 import BeautifulSoup
from xml.dom import minidom
from collections import defaultdict
from openpyxl import load_workbook
from xml.etree import ElementTree
import glob, os
import os.path
import datetime


routes=[]

def init_routes():
    if len(routes)>0:
        return
    loaded_wb = load_workbook(filename=os.path.dirname(os.path.realpath(__file__))+"/result.xlsx")
    sheet = loaded_wb.worksheets[0]

    for i in range(1,sheet.max_row+1):
        route=[]
        col=1
        while(sheet.cell(row=i,column=col).value):
            route.append(sheet.cell(row=i,column=col).value)
            col = col + 1
        routes.append(route)

def find_route(query):
    global routes
    res = []
    for route in routes:
        if query in route:
            res.append(route)
    return res

def find_route_with_queries(queries):
    global routes
    result = routes
    for i in range(len(queries)):
        result = find_route(queries[i])
    return result

def filter_by_activeness(activeness, routes, days):
    indexes = []
    for route in routes:
        indexes.append(len(route)/days*10)
    sorted_route = {}

    res = []
    for i,route in enumerate(routes):
        if int(indexes[i]) in range(0,2):
            res.append(route)
    sorted_route[1]=res
    res = []
    for i,route in enumerate(routes):
        if int(indexes[i]) in range(3,4):
            res.append(route)
    sorted_route[2]=res
    res = []
    for i,route in enumerate(routes):
        if int(indexes[i]) in range(5,6):
            res.append(route)
    sorted_route[3]=res
    res = []
    for i,route in enumerate(routes):
        if int(indexes[i]) in range(7,9):
            res.append(route)
    sorted_route[4]=res
    res = []
    for i,route in enumerate(routes):
        if int(indexes[i]) in range(10,100):
            res.append(route)
    sorted_route[5]=res

    if activeness in sorted_route:
        return sorted_route[activeness]
    else:
        candidate=0
        minimum=100
        for i,a in enumerate(sorted_route):
            if len(sorted_route[i+1])>0:
                if minimum > abs(i+1 - int(activeness)):
                    minimum = abs(i+1 - int(activeness))
                    candidate = i+1
        print(str(sorted_route))
        if minimum!=0:
            return sorted_route[candidate]
        else:
            return None

def routefinder(city_list, activeness, days):
    init_routes()
    return filter_by_activeness(activeness,find_route_with_queries(city_list),days)


class Airport:
    def __init__(self, name, code, region, country):
        self.name = name
        self.code = code
        self.region = region
        self.country = country
    @property
    def getName(self):
        return self.name

    @property
    def getCode(self):
        return self.code

    @property
    def getRegion(self):
        return self.code

    @property
    def getCountry(self):
        return self.country

def getUrl(s, d, date, people_number):
    return "https://www.google.com/flights/?f=0&gl#search;f="+quote(str(s)).replace("%27","").replace("%5B","").replace("%5D","")+";t="+quote(str(d)).replace("%27","").replace("%5B","").replace("%5D","")+";d="+date+";tt=o"+";px="+people_number

def getCityData(query):
    try:
        url = 'https://maps.googleapis.com/maps/api/geocode/xml?address='+quote(query)
        request = Request(url)
        request.add_header('Accept-Encoding', 'utf-8')

        response = urlopen(request)
        tree = ElementTree.parse(response)

        response_tag = tree.getroot()
        result = response_tag.find("result")

        for a in result.findall("address_component"):
            if a.find("type").text == "country":
                country = a.find("long_name").text
        city_name=result.find("address_component").find("long_name").text

        lat = result.find("geometry").find("location").find("lat").text
        lng = result.find("geometry").find("location").find("lng").text
        print([city_name, country, lat, lng])
        return [city_name, country, lat, lng]
    except:
        return None

def getAirport(query):
    if query is None:
        return None
    f = codecs.open(os.path.dirname(os.path.realpath(__file__))+"/airports.dat","r","utf-8")
    lines = f.readlines()
    airport_info = defaultdict(list)
    for line in lines:
        temp = quote(line).split('%2C')

        name = unquote(temp[1].replace("%22","").replace("%20"," "))
        code = temp[4].replace("%22","")
        region = unquote(temp[2].replace("%22","").replace("%20"," "))
        country = temp[3].replace("%22","").replace("%20"," ")

        if code != "%5CN":
            airport_info[region].append(Airport(name,code,region,country))
    f.close()
    if len(airport_info[query]) != 0:
        return airport_info[query]
    else:
        if 'City' in query:
            return getAirport(query.split(" ")[0])

        # Try again search airport names
        for line in lines:
            temp = quote(line).split('%2C')

            name = unquote(temp[1].replace("%22","").replace("%20"," "))
            code = temp[4].replace("%22","")
            region = unquote(temp[2].replace("%22","").replace("%20"," "))
            country = temp[3].replace("%22","").replace("%20"," ")

            if query in name:
                return airport_info[region]

        return None

def getMinimumFare(s, d, date, people_number):
    api_key = "AIzaSyD3aOnfRprdxCdxJuBaZgiVKibwdgHmnKU"
    url = "https://www.googleapis.com/qpxExpress/v1/trips/search?key=" + api_key
    headers = {'content-type': 'application/json'}

    values = {
        "request": {
        "passengers": {
          "kind": "qpxexpress#passengerCounts",
          "adultCount": people_number,
        },
        "slice": [
          {
            "kind": "qpxexpress#sliceInput",
            "origin": s,
            "destination": d,
            "date": date,
          }
        ],
        "saleCountry": "KR",
        "refundable": "false",
        "solutions": 1
      }
    }
    data = json.dumps(values)
    data = data.encode("utf-8")

    req = Request(url, data=data, headers = headers)
    f = urlopen(req)
    response = f.read()
    result = json.loads(response.decode("utf-8"))
    fare = result['trips']['tripOption'][0]['saleTotal']
    currency = fare[0:3]
    fare = fare[3:len(fare)]
    return [fare,currency]

def index(request):
    return render(request, 'routesearch/index.html', {})

def route(request):
    cities = []

    if 'user_name' in request.GET:
        user_name = request.GET['user_name']
    if 'city' in request.GET:
        cities = request.GET.getlist('city')
    if 'start_date' in request.GET:
        start_date = request.GET['start_date']
    if 'end_date' in request.GET:
        end_date = request.GET['end_date']
    if 'people' in request.GET:
        people_number = request.GET['people']
    if 'style' in request.GET:
        style = request.GET['style']
    route_list = routefinder(cities,style,(datetime.datetime.strptime(end_date,'%Y-%m-%d')-datetime.datetime.strptime(start_date,'%Y-%m-%d')).days)
    return render_to_response('routesearch/show_routes.html',{'user_name':user_name,
                                                        'route':route,
                                                        'start_date':start_date,
                                                        'end_date':end_date,
                                                        'people_number':people_number,
                                                        'route_list':route_list})

def result(request):
    cities = []
    route = []
    route_coordinates=[]
    countries = []
    omitted = []

    if 'user_name' in request.GET:
        user_name = request.GET['user_name']
    if 'city' in request.GET:
        cities = request.GET.getlist('city')
    if 'start_date' in request.GET:
        start_date = request.GET['start_date']
    if 'end_date' in request.GET:
        end_date = request.GET['end_date']
    if 'people' in request.GET:
        people_number = request.GET['people']
    if 'style' in request.GET:
        style = request.GET['style']


    for city in cities:
        if city is None:
            continue
        try:
            temp = getCityData(city)
            route.append(temp[0])
            route_coordinates.append([temp[2],temp[3]])
            if getCityData(city)[1] not in countries:
                    countries.append(temp[1])
        except:
            omitted.append(city)

    if len(cities)<1:
        if 'code1' in request.GET:
            sf = request.GET['code1']
        if 'code2' in request.GET:
            st = request.GET.getlist('code2')
        if 'code3' in request.GET:
            ef = request.GET['code3']
        if 'code4' in request.GET:
            et = request.GET['code4']
        if 'start_parameter' in request.GET:
            start_date = request.GET['start_parameter']
        if 'end_parameter' in request.GET:
            end_date = request.GET['end_parameter']
        if 'people_parameter' in request.GET:
            pn = request.GET['people_parameter']

        fare1 = getMinimumFare(sf,st,start_date,pn)
        fare2 = getMinimumFare(ef,et,end_date,pn)
        currency1 = fare1[1]
        currency2 = fare2[1]

        start = getUrl(sf,st,start_date,pn)
        end = getUrl(ef,et,end_date,pn)

        return render_to_response('routesearch/fares.html',{'fare1':fare1[0],
                                                            'fare2':fare2[0],
                                                            'currency1':currency1,
                                                            'currency2':currency2,
                                                            'start':start,
                                                            'end':end,
                                                            'total':int(fare1[0])+int(fare2[0])})

    c1 = cities[0]
    c2 = cities[len(cities)-1]

    if (c1 and c2) not in omitted:
        airport1 = getAirport(getCityData(c1)[0])
        airport2 = getAirport(getCityData(c2)[0])


    return render_to_response('routesearch/result.html',{'user_name':user_name,
                                                        'route':route,
                                                        'route_coordinates':route_coordinates,
                                                        'start_date':start_date,
                                                        'end_date':end_date,
                                                        'people_number':people_number,
                                                        'airport1':airport1,
                                                        'airport2':airport2,
                                                        'airport3':getAirport('Seoul')+getAirport('Busan'),
                                                        'airport4':getAirport('Seoul')+getAirport('Busan'),})


def fares(request):
    return render(request,'routesearch/empty.html')
